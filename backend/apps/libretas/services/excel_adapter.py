# backend/apps/libretas/services/excel_adapter.py
from __future__ import annotations
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, Dict, Any
from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook

@dataclass(frozen=True)
class ExcelBaseInfo:
    path: Path
    filename: str
    sheetnames: tuple[str, ...]

def leer_base(path: str) -> ExcelBaseInfo:
    """
    Carga el workbook y devuelve datos base.
    """
    p = Path(path)
    wb = load_workbook(p, data_only=True)
    return ExcelBaseInfo(path=p, filename=p.name, sheetnames=tuple(wb.sheetnames))

def iter_alumnos_hoja(wb: Workbook, sheet_name: str) -> Iterable[Dict[str, Any]]:
    """
    Convención dummy para pruebas/UGEL:
      A: alumnoId, B: alumno, C-F: B1..B4, G: curso (opcional)
      Encabezados en fila 1, datos desde fila 2.
    """
    ws = wb[sheet_name]
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[0] is None:
            continue
        yield {
            "alumnoId": r[0],
            "alumno": r[1],
            "B1": r[2],
            "B2": r[3],
            "B3": r[4],
            "B4": r[5],
            "curso": r[6] if len(r) > 6 else None,
        }

def escribir_rangos_consolidado(wb: Workbook, sheet_name: str, rows: Iterable[Dict[str, Any]]) -> None:
    """
    Escribe solo promedio/letra/comentario sin romper formato:
      H: promedio (num), I: letra (texto), J: comentario (texto)
      Empareja por alumnoId (col A).
    """
    ws = wb[sheet_name]
    idx_por_id: Dict[Any, int] = {}
    for row in range(2, ws.max_row + 1):
        alumno_id = ws.cell(row=row, column=1).value
        if alumno_id is not None:
            idx_por_id[alumno_id] = row

    for item in rows:
        rid = idx_por_id.get(item.get("alumnoId"))
        if not rid:
            continue
        if "promedio" in item:
            ws.cell(row=rid, column=8, value=float(item["promedio"]))  # H
        if "letra" in item:
            ws.cell(row=rid, column=9, value=str(item["letra"]))       # I
        if "comentario" in item:
            ws.cell(row=rid, column=10, value=str(item["comentario"])) # J

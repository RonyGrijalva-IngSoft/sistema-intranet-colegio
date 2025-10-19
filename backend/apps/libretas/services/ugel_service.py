# backend/apps/libretas/services/ugel_service.py
import os
import uuid
from typing import List, Tuple
from django.conf import settings

from openpyxl import load_workbook

UPLOADS_DIR = os.path.join(settings.MEDIA_ROOT, "ugel", "uploads")
EXPORTS_DIR = os.path.join(settings.MEDIA_ROOT, "ugel", "exports")

def _ensure_dirs():
    os.makedirs(UPLOADS_DIR, exist_ok=True)
    os.makedirs(EXPORTS_DIR, exist_ok=True)

def guardar_upload(django_file) -> Tuple[str, str]:
    """
    Guarda .xlsx preservando el nombre original, bajo media/ugel/uploads/<uuid>/<filename>.
    Retorna (upload_id, ruta_absoluta_guardada)
    """
    _ensure_dirs()
    original_name = django_file.name
    upload_id = str(uuid.uuid4())
    target_dir = os.path.join(UPLOADS_DIR, upload_id)
    os.makedirs(target_dir, exist_ok=True)

    target_path = os.path.join(target_dir, original_name)
    with open(target_path, "wb") as out:
        for chunk in django_file.chunks():
            out.write(chunk)
    return upload_id, target_path

def inspeccionar_hojas(path: str) -> List[str]:
    """
    Retorna la lista de hojas del Excel subido (para UI).
    """
    wb = load_workbook(path, data_only=True, read_only=True)
    return wb.sheetnames

# backend/apps/libretas/tests/test_ugel.py
from django.test import TestCase
from django.contrib.auth import get_user_model
from rest_framework.test import APIClient
from django.core.files.uploadedfile import SimpleUploadedFile
from openpyxl import Workbook
from io import BytesIO

User = get_user_model()

class UgelUploadTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(username="tester", password="pass")
        self.client = APIClient()
        self.client.login(username="tester", password="pass")

    def _make_xlsx(self) -> SimpleUploadedFile:
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Resumen"
        ws1.append(["Alumno", "Nota"])
        ws1.append(["A", 14.5])

        wb.create_sheet("Matemática")
        wb.create_sheet("Comunicación")

        stream = BytesIO()
        wb.save(stream)
        stream.seek(0)
        return SimpleUploadedFile("UGEL_dummy.xlsx", stream.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    def test_upload_y_hojas(self):
        f = self._make_xlsx()
        resp = self.client.post("/libretas/ugel/upload", {"file": f}, format="multipart")
        self.assertEqual(resp.status_code, 201)
        data = resp.json()
        self.assertIn("uploadId", data)
        self.assertEqual(data["filename"], "UGEL_dummy.xlsx")
        self.assertIn("Resumen", data["hojas"])

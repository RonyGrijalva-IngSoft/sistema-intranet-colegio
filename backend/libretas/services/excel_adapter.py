from openpyxl import load_workbook
import unicodedata

def _norm(s):
    if s is None:
        return ""
    s = str(s)
    s = "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")
    return s.strip().upper().replace(" ", "_")

def _s(s):
    """string seguro (evita .strip sobre no-str)"""
    return str(s).strip() if s is not None else ""

def _f(x):
    """float seguro (convierte celdas vacías o texto no numérico a 0.0)"""
    try:
        if x is None or (isinstance(x, str) and x.strip() == ""):
            return 0.0
        return float(x)
    except Exception:
        return 0.0

class ExcelAdapter:
    """
    Lee un Excel local y expone:
      - alumnos_por_seccion(seccion_id)
      - cursos_de_grado(grado)
      - notas_de_curso(grado, seccion_id, curso_id)

    Cabeceras esperadas (detecta sinónimos):
      ID | APELLIDOS | NOMBRES | CURSO_ID | B1 | B2 | B3 | B4
    """

    _SYN = {
        "id": {"ID", "CODIGO", "ALUMNO_ID", "ID_ALUMNO", "DNI"},
        "apellidos": {"APELLIDOS", "APELLIDO", "APE"},
        "nombres": {"NOMBRES", "NOMBRE", "NOM"},
        "curso_id": {"CURSO_ID", "ID_CURSO", "COD_CURSO", "CURSO", "AREA_ID"},
        "b1": {"B1", "BIM1", "BIMESTRE1"},
        "b2": {"B2", "BIM2", "BIMESTRE2"},
        "b3": {"B3", "BIM3", "BIMESTRE3"},
        "b4": {"B4", "BIM4", "BIMESTRE4"},
    }

    def __init__(self, ruta_excel, hoja=None, cols_override=None):
        self.ruta_excel = ruta_excel
        self.hoja = hoja
        self.cols_override = {k: _norm(v) for k, v in (cols_override or {}).items()}

    def _open(self):
        wb = load_workbook(self.ruta_excel, data_only=True)
        ws = wb[self.hoja] if (self.hoja and self.hoja in wb.sheetnames) else wb[wb.sheetnames[0]]
        return ws

    def _header_index(self, ws):
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        norm = [_norm(h) for h in headers]
        idx = {norm[i]: i for i in range(len(norm))}
        return idx, norm

    def _map_columns(self, idx, headers_norm):
        if self.cols_override:
            missing = [v for v in self.cols_override.values() if v not in idx]
            if missing:
                raise ValueError(f"Columnas override no encontradas: {missing}\nCabeceras: {headers_norm}")
            return self.cols_override
        mapping = {}
        for key, syns in self._SYN.items():
            found = None
            for s in syns:
                n = _norm(s)
                if n in idx:
                    found = n
                    break
            if not found and key in {"b1","b2","b3","b4"}:
                continue  # si no está, se toma como 0
            if not found:
                raise ValueError(f"No se detectó la columna requerida '{key}'. Cabeceras: {headers_norm}")
            mapping[key] = found
        return mapping

    def _rows(self):
        ws = self._open()
        idx, headers_norm = self._header_index(ws)
        cols = self._map_columns(idx, headers_norm)
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            def val(col_key, default=None):
                cname = cols.get(col_key)
                if not cname or cname not in idx:
                    return default
                j = idx[cname]
                return row[j]
            rows.append({
                "id": val("id"),
                "apellidos": _s(val("apellidos", "")),
                "nombres": _s(val("nombres", "")),
                "curso_id": val("curso_id"),
                "B1": _f(val("b1", 0)),
                "B2": _f(val("b2", 0)),
                "B3": _f(val("b3", 0)),
                "B4": _f(val("b4", 0)),
            })
        return rows

    # ==== Métodos IDataPort ====
    def alumnos_por_seccion(self, seccion_id):
        rows = self._rows()
        alumnos = {}
        for r in rows:
            if r["id"] is None:
                continue
            aid = int(_f(r["id"]))
            if aid not in alumnos:
                alumnos[aid] = {"id": aid, "apellidos": r["apellidos"], "nombres": r["nombres"]}
        return list(alumnos.values())

    def cursos_de_grado(self, grado):
        rows = self._rows()
        cursos = {}
        for r in rows:
            if r["curso_id"] is None:
                continue
            cid = int(_f(r["curso_id"]))
            if cid not in cursos:
                cursos[cid] = {"id": cid, "nombre": f"Curso {cid}"}
        return list(cursos.values())

    def notas_de_curso(self, grado, seccion_id, curso_id):
        rows = self._rows()
        out = {}
        for r in rows:
            if r["curso_id"] is None:
                continue
            if int(_f(r["curso_id"])) != int(_f(curso_id)):
                continue
            if r["id"] is None:
                continue
            aid = int(_f(r["id"]))
            out[aid] = {"B1": r["B1"], "B2": r["B2"], "B3": r["B3"], "B4": r["B4"]}
        return out
